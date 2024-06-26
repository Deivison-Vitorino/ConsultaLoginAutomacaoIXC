# manipulador_excel.py

import openpyxl

class ManipuladorExcel:
    def __init__(self, arquivo_entrada: str, arquivo_saida: str):
        self.arquivo_entrada = arquivo_entrada
        self.arquivo_saida = arquivo_saida
        self.wb_entrada = openpyxl.load_workbook(arquivo_entrada)
        self.wb_saida = openpyxl.load_workbook(arquivo_saida)
        self.sheet_entrada = self.wb_entrada['logins']
        self.sheet_saida = self.wb_saida['PortaSinalBairroCLiente']
        
    def ler_logins(self):
        """
        Lê os logins da planilha de entrada.

        :return: Uma lista de tuplas contendo login, porta e potência.
        """
        return [linha for linha in self.sheet_entrada.iter_rows(min_row=2, values_only=True)]

    def salvar_dados_cliente(self, login, porta, potencia, nome, endereco):
        """
        Salva os dados do cliente na planilha de saída.

        :paramentro login: O login do cliente.
        :paramentro porta: A porta da OLT cujo cliente está conectado.
        :paramentro potencia: A potência de sinal(dBm) da fibra do cliente.
        :paramentro nome: O nome do cliente.
        :paramentro endereco: O endereço do cliente.
        """
        self.sheet_saida.append([login, porta, potencia, endereco, nome])
        self.wb_saida.save(self.arquivo_saida)
